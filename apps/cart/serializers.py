
from django.contrib.auth import get_user_model

from rest_framework import serializers

from apps.cart.models import (
    Order, OrderItem
)

from apps.products.models import Product

from openpyxl import Workbook

from apps.cart.send_mail import send_msg
from time import sleep

User = get_user_model()


class UserSerializer(serializers.ModelSerializer):
    class Meta:
        model = User
        fields = ('__all__')


class OrderItemSerializer(serializers.ModelSerializer):
    class Meta:
        model = OrderItem
        fields = ["order", 'product', 'amount']


class OrderSerializer(serializers.ModelSerializer):
    products = serializers.ListField(write_only=True)

    class Meta:
        model = Order
        fields = (
            'id', 'email', 'addres', 'full_name',
            'phone', 'start_date', 'ordered',
            'products',
        )
#  [{'prod_id': 1, 'amount': 10}, {'prod_id': 2, 'amount': 3}]

    def create(self, validated_data):
        products = validated_data.get('products')
        email=validated_data.get('email')
        order = Order.objects.create(
            email=email,
            addres=validated_data.get('addres'),
            full_name=validated_data.get('full_name'),
            phone=validated_data.get('phone'),
            ordered=validated_data.get('ordered')
        )
        wb = Workbook()
        ws = wb.active
        lineA = 'A'
        lineB = 'B'
        lineC = 'C'
        lineD = 'D'
        ws[lineA+'1'] = 'Название'
        ws[lineB+'1'] = 'Количество'
        ws[lineC+'1'] = 'Цена'
        ws[lineD+'1'] = 'Общая цена'
        for p in products:
            product = Product.objects.filter(id=p['prod_id']).first()
            OrderItem.objects.create(
                order=order,
                product=product,
                amount=p['amount']
            )
        line = 0
        all_price = 0
        for i in range(len(products)):
            line = i+2
            sl = str(line)
            product = Product.objects.filter(id=p['prod_id']).first()
            all_price += product.price*products[i]['amount']
            ws[lineA+sl] = product.name
            ws[lineB+sl] = products[i]['amount']
            ws[lineC+sl] = product.price
            ws[lineD+sl] = product.price*products[i]['amount']
        line+=1
        ws[lineD+str(line)] = all_price
        wb.save(f"orders/order{order.id}.xlsx")
        send_msg(file=f"orders/order{order.id}.xlsx", email=email)
        return order




# {
#     "email": "lev201611@gmail.com",
#     "addres": "logvinenko 55A",
#     "full_name": "lev Boiko",
#     "phone": "0559595139",
#     "ordered": false,
#     "products": [{"prod_id": 1, "amount": 10}, {"prod_id": 2, "amount": 3}]
# }